import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.services import dashboard_service

HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
BODY_FONT = Font(name="Arial")


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.font = BODY_FONT

    for col_idx, header in enumerate(headers, start=1):
        cell_lengths = [len(str(header))] + [len(str(r[col_idx - 1])) for r in rows]
        max_len = max(cell_lengths)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 4, 50))


def generate_operations_report_xlsx(db: Session) -> bytes:
    wb = Workbook()

    # --- Summary sheet -----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary = dashboard_service.get_summary(db)
    _write_sheet(
        ws_summary,
        ["Metric", "Value"],
        [
            ["Total Customers", summary.total_customers],
            ["Total Packages", summary.total_packages],
            ["Active Tours", summary.active_tours],
            ["Total Bookings", summary.total_bookings],
            ["Confirmed Bookings", summary.confirmed_bookings],
            ["Cancelled Bookings", summary.cancelled_bookings],
            ["Total Revenue", summary.total_revenue],
            ["Total Refunds", summary.total_refunds],
            ["Average Package Rating", summary.average_package_rating],
        ],
    )

    # --- Popular destinations ------------------------------------------------
    ws_dest = wb.create_sheet("Popular Destinations")
    dests = dashboard_service.get_popular_destinations(db, limit=20)
    _write_sheet(
        ws_dest,
        ["Destination ID", "Destination", "Booking Count"],
        [[d.destination_id, d.destination_name, d.booking_count] for d in dests],
    )

    # --- Popular packages -----------------------------------------------------
    ws_pkg = wb.create_sheet("Popular Packages")
    pkgs = dashboard_service.get_popular_packages(db, limit=20)
    _write_sheet(
        ws_pkg,
        ["Package ID", "Package", "Booking Count"],
        [[p.package_id, p.package_name, p.booking_count] for p in pkgs],
    )

    # --- Package performance -----------------------------------------------------
    ws_perf = wb.create_sheet("Package Performance")
    perf = dashboard_service.package_performance_report(db)
    _write_sheet(
        ws_perf,
        ["Package ID", "Package", "Total Bookings", "Total Revenue"],
        [[r["package_id"], r["package_name"], r["total_bookings"], r["total_revenue"]] for r in perf],
    )

    # --- Destination revenue -----------------------------------------------------
    ws_rev = wb.create_sheet("Destination Revenue")
    dest_rev = dashboard_service.destination_wise_revenue(db)
    _write_sheet(
        ws_rev,
        ["Destination ID", "Destination", "Revenue"],
        [[r["destination_id"], r["destination_name"], r["revenue"]] for r in dest_rev],
    )

    # --- Cancellations -----------------------------------------------------------
    ws_cancel = wb.create_sheet("Cancellations")
    cancel_report = dashboard_service.cancellation_report(db)
    _write_sheet(
        ws_cancel,
        ["Metric", "Value"],
        [
            ["Total Cancellations", cancel_report["total_cancellations"]],
            ["Total Refunded", cancel_report["total_refunded"]],
        ],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
