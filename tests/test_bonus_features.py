from openpyxl import load_workbook
import io


def _setup_confirmed_booking(client, headers):
    dest = client.post("/destinations", json={"name": "Goa", "country": "India"}, headers=headers).json()
    pkg = client.post(
        "/packages",
        json={
            "package_name": "Goa Trip",
            "destination_id": dest["id"],
            "duration_days": 5,
            "base_price": 5000,
            "max_capacity": 10,
            "start_date": "2026-12-01",
            "end_date": "2026-12-06",
        },
        headers=headers,
    ).json()
    client.put(f"/packages/{pkg['id']}", json={"status": "Published"}, headers=headers)
    cust = client.post(
        "/customers", json={"name": "Bonus Cust", "email": "bonuscust@test.com", "phone": "9998887777"}, headers=headers
    ).json()
    booking = client.post(
        "/bookings",
        json={"customer_id": cust["id"], "package_id": pkg["id"], "number_of_travelers": 1},
        headers=headers,
    ).json()
    client.post(
        f"/payments/{booking['id']}",
        json={"amount": booking["total_amount"], "payment_method": "UPI", "transaction_id": "TXN-BONUS-1"},
        headers=headers,
    )
    return booking


def test_versioned_api_prefix_serves_same_data(client, admin_auth_headers):
    unversioned = client.get("/auth/me", headers=admin_auth_headers)
    versioned = client.get("/api/v1/auth/me", headers=admin_auth_headers)
    assert unversioned.status_code == versioned.status_code == 200
    assert unversioned.json() == versioned.json()


def test_booking_confirmation_pdf_download(client, admin_auth_headers):
    booking = _setup_confirmed_booking(client, admin_auth_headers)
    resp = client.get(f"/bookings/{booking['id']}/confirmation-pdf", headers=admin_auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"
    assert len(resp.content) > 500


def test_booking_qr_code_download(client, admin_auth_headers):
    booking = _setup_confirmed_booking(client, admin_auth_headers)
    resp = client.get(f"/bookings/{booking['id']}/qr-code", headers=admin_auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "image/png"
    assert resp.content[:8] == b"\x89PNG\r\n\x1a\n"  # PNG magic bytes


def test_confirmation_pdf_404_for_missing_booking(client, admin_auth_headers):
    resp = client.get("/bookings/999999/confirmation-pdf", headers=admin_auth_headers)
    assert resp.status_code == 404


def test_excel_operations_report_export(client, admin_auth_headers):
    _setup_confirmed_booking(client, admin_auth_headers)
    resp = client.get("/dashboard/reports/export", headers=admin_auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(io.BytesIO(resp.content))
    assert "Summary" in wb.sheetnames
    assert "Popular Destinations" in wb.sheetnames
    summary_rows = list(wb["Summary"].iter_rows(values_only=True))
    assert summary_rows[0] == ("Metric", "Value")
    metrics = {row[0]: row[1] for row in summary_rows[1:]}
    assert metrics["Total Bookings"] == 1
    assert metrics["Confirmed Bookings"] == 1


def test_dashboard_summary_cache_returns_consistent_values(client, admin_auth_headers):
    _setup_confirmed_booking(client, admin_auth_headers)
    first = client.get("/dashboard/summary", headers=admin_auth_headers).json()
    second = client.get("/dashboard/summary", headers=admin_auth_headers).json()
    assert first == second
    assert first["total_bookings"] == 1


def test_celery_task_runs_eagerly_without_broker(client, admin_auth_headers, db_session):
    """The Celery scaffold defaults to an in-process eager transport when no
    CELERY_BROKER_URL is configured, so the task body itself is directly
    testable here without standing up Redis/RabbitMQ."""
    from app.tasks import send_upcoming_tour_reminders

    result = send_upcoming_tour_reminders.delay(3)
    # Runs inline (task_always_eager=True); .get() returns the task's return value.
    assert isinstance(result.get(timeout=5), int)
